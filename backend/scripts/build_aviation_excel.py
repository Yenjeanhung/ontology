"""从种子数据生成「民航维修领域本体」分菜单 Excel 文件。

不操作数据库，仅读取 seed_aviation_ontology.py 中的数据定义并写出 xlsx。
按菜单拆分为 4 个文件，对应 4 个子页面：
  - 本体管理.xlsx      （本体类别 / 本体 / 属性 / 模板绑定）
  - 关系字典.xlsx      （关系）
  - 本体关系.xlsx      （约束）
  - 本体模板.xlsx      （属性模板 / 模板属性）

用法：
    cd backend
    python scripts/build_aviation_excel.py
"""

from __future__ import annotations

import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).parent.parent))
sys.path.insert(0, str(Path(__file__).parent))

from openpyxl import Workbook

from seed_aviation_ontology import (
    ATTRIBUTES,
    ATTRIBUTE_TEMPLATES,
    COLORS,
    CONSTRAINTS,
    CATEGORY_DESC,
    CATEGORY_NAME,
    ONTOLOGIES,
    RELATIONS,
    TEMPLATE_BINDINGS,
)
from services.ontology_import_service import (
    HEADERS,
    SCOPE_SHEETS,
    _write_sheet,
)


def _build_full_workbook() -> Workbook:
    wb = Workbook()
    wb.remove(wb.active)

    # 本体类别
    _write_sheet(wb, "本体类别", HEADERS["本体类别"], [[CATEGORY_NAME, CATEGORY_DESC, "否"]])

    # 本体
    ont_rows = []
    for i, o in enumerate(ONTOLOGIES, start=1):
        ont_rows.append([CATEGORY_NAME, o["name"], o["group"], o["description"], COLORS[o["group"]], i])
    _write_sheet(wb, "本体", HEADERS["本体"], ont_rows)

    # 属性
    attr_rows = []
    for ont_name, attrs in ATTRIBUTES.items():
        for i, a in enumerate(attrs, start=1):
            attr_rows.append([
                CATEGORY_NAME, ont_name, a["name"], a.get("code", ""), a["data_type"],
                "是" if a.get("is_required") else "否", a.get("description", ""),
                a.get("default_value", ""), i,
            ])
    _write_sheet(wb, "属性", HEADERS["属性"], attr_rows)

    # 关系
    rel_rows = [[CATEGORY_NAME, r["name"], "", r["description"]] for r in RELATIONS]
    _write_sheet(wb, "关系", HEADERS["关系"], rel_rows)

    # 约束
    con_rows = [[CATEGORY_NAME, s, r, t, ""] for s, r, t in CONSTRAINTS]
    _write_sheet(wb, "约束", HEADERS["约束"], con_rows)

    # 属性模板
    tpl_rows = [[t["name"], t["description"]] for t in ATTRIBUTE_TEMPLATES]
    _write_sheet(wb, "属性模板", HEADERS["属性模板"], tpl_rows)

    # 模板属性
    tpl_attr_rows = []
    for t in ATTRIBUTE_TEMPLATES:
        for i, a in enumerate(t["attributes"], start=1):
            tpl_attr_rows.append([
                t["name"], a["name"], a.get("code", ""), a["data_type"],
                "是" if a.get("is_required") else "否", a.get("description", ""),
                a.get("default_value", ""), i,
            ])
    _write_sheet(wb, "模板属性", HEADERS["模板属性"], tpl_attr_rows)

    # 模板绑定
    bind_rows = []
    for ont_name, tpls in TEMPLATE_BINDINGS.items():
        for t in tpls:
            bind_rows.append([CATEGORY_NAME, ont_name, t])
    _write_sheet(wb, "模板绑定", HEADERS["模板绑定"], bind_rows)

    return wb


def build() -> list[Path]:
    full_wb = _build_full_workbook()
    out_dir = Path(__file__).parent.parent / "data"
    out_dir.mkdir(parents=True, exist_ok=True)

    scope_names = {
        "ontologies": "本体管理",
        "relations": "关系字典",
        "constraints": "本体关系",
        "templates": "本体模板",
    }

    paths: list[Path] = []
    # 同时保留一个完整版（方便高级用户一次性导入）
    full_path = out_dir / "民航维修领域本体-完整版.xlsx"
    full_wb.save(full_path)
    paths.append(full_path)

    for scope, label in scope_names.items():
        scoped_wb = Workbook()
        scoped_wb.remove(scoped_wb.active)
        keep_sheets = set(SCOPE_SHEETS[scope])
        for sheet_name in full_wb.sheetnames:
            if sheet_name in keep_sheets:
                # copy worksheet to new workbook
                src = full_wb[sheet_name]
                dst = scoped_wb.create_sheet(title=sheet_name)
                for row in src.iter_rows(values_only=True):
                    dst.append(row)
                # 复制列宽
                for col_letter, dim in src.column_dimensions.items():
                    if dim.width:
                        dst.column_dimensions[col_letter].width = dim.width
                dst.freeze_panes = src.freeze_panes
        path = out_dir / f"民航维修领域本体-{label}.xlsx"
        scoped_wb.save(path)
        paths.append(path)

    return paths


if __name__ == "__main__":
    paths = build()
    for p in paths:
        print(f"已生成：{p}  ({p.stat().st_size} 字节)")
