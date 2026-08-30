"""本体 Excel 导入 / 导出服务。

设计要点
--------
1. **逐行校验**：非法行跳过（不插入）并记录失败原因；合法行继续导入。
2. **重复覆盖**：按业务唯一键判断，已存在则更新，不存在则新建。
3. **不做删除**：文件中缺失的既有数据不会被删除，只做新增 + 覆盖，避免误删。
4. **完整报告**：返回总数 / 成功 / 失败 / 新建 / 更新，以及逐行失败原因（含 sheet 名与行号）。

Excel 结构（8 个 sheet，首行为表头；`*` 表示必填）
------------------------------------------------
| sheet      | 列 |
|---|---|
| 本体类别   | 类别名称* | 描述 | 是否系统内置 |
| 本体       | 类别名称 | 本体名称* | 分组 | 描述 | 颜色 | 排序 |
| 属性       | 类别名称 | 本体名称* | 属性名* | 编码 | 数据类型* | 必填 | 描述 | 默认值 | 排序 |
| 关系       | 类别名称 | 关系名称* | 编码 | 描述 |
| 约束       | 类别名称 | 起点本体* | 关系* | 终点本体* | 描述 |
| 属性模板   | 模板名* | 模板描述 |
| 模板属性   | 模板名* | 属性名* | 编码 | 数据类型* | 必填 | 描述 | 默认值 | 排序 |
| 模板绑定   | 类别名称 | 本体名称* | 模板名* |

「类别名称」列在子表中可选：
- 文件中仅定义了一个类别时，子表可省略该列（自动归属该类别）；
- 定义了多个类别时，子表必须显式指定，否则该行报错。

「分组」列仅用于在本表未填颜色时自动分配配色，不落库（`ontologies` 表无此字段）。
"""

from __future__ import annotations

import io
import json
import re
import uuid
from datetime import datetime
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from sqlalchemy import delete, select
from sqlalchemy.ext.asyncio import AsyncSession

from models import (
    Ontology,
    OntologyAttribute,
    OntologyAttributeTemplate,
    OntologyCategory,
    OntologyRelation,
    OntologyRelationConstraint,
    OntologyTemplateAttribute,
    OntologyTemplateBinding,
)

# ─────────────────────────── 常量定义 ───────────────────────────

SHEET_CATEGORY = "本体类别"
SHEET_ONTOLOGY = "本体"
SHEET_ATTRIBUTE = "属性"
SHEET_RELATION = "关系"
SHEET_CONSTRAINT = "约束"
SHEET_TEMPLATE = "属性模板"
SHEET_TEMPLATE_ATTR = "模板属性"
SHEET_BINDING = "模板绑定"

HEADERS: dict[str, list[str]] = {
    SHEET_CATEGORY: ["类别名称*", "描述", "是否系统内置"],
    SHEET_ONTOLOGY: ["类别名称", "本体名称*", "分组", "描述", "颜色", "排序"],
    SHEET_ATTRIBUTE: ["类别名称", "本体名称*", "属性名*", "编码", "数据类型*", "必填", "描述", "默认值", "排序"],
    SHEET_RELATION: ["类别名称", "关系名称*", "编码", "描述"],
    SHEET_CONSTRAINT: ["类别名称", "起点本体*", "关系*", "终点本体*", "描述"],
    SHEET_TEMPLATE: ["模板名*", "模板描述"],
    SHEET_TEMPLATE_ATTR: ["模板名*", "属性名*", "编码", "数据类型*", "必填", "描述", "默认值", "排序"],
    SHEET_BINDING: ["类别名称", "本体名称*", "模板名*"],
}

# scope -> 该 scope 下参与导入/导出的 sheet 集合
# full: 全部；ontologies: 本体管理页；relations: 关系字典；constraints: 本体关系；templates: 本体模板
VALID_SCOPES = {"full", "ontologies", "relations", "constraints", "templates"}
SCOPE_SHEETS: dict[str, list[str]] = {
    "full": [SHEET_CATEGORY, SHEET_ONTOLOGY, SHEET_ATTRIBUTE, SHEET_RELATION,
             SHEET_CONSTRAINT, SHEET_TEMPLATE, SHEET_TEMPLATE_ATTR, SHEET_BINDING],
    "ontologies": [SHEET_CATEGORY, SHEET_ONTOLOGY, SHEET_ATTRIBUTE, SHEET_BINDING],
    "relations": [SHEET_RELATION],
    "constraints": [SHEET_CONSTRAINT],
    "templates": [SHEET_TEMPLATE, SHEET_TEMPLATE_ATTR],
}

# 系统支持的数据类型（与前端下拉一致）
DATA_TYPES = {"string", "text", "number", "boolean", "date", "datetime"}

# 分组 → 默认配色（本表未指定「颜色」时使用）
GROUP_COLORS = {
    "航空器结构类": "#5470c6",
    "故障类": "#ee6666",
    "维修类": "#91cc75",
    "适航文件类": "#fac858",
    "设备结构类": "#5470c6",
    "资源类": "#73c0de",
}
PALETTE = ["#5470c6", "#91cc75", "#fac858", "#ee6666", "#73c0de", "#3ba272",
           "#fc8452", "#9a60b4", "#ea7ccc"]

# 长度约束（与 schema.sql 保持一致）
LEN_NAME = 50
LEN_DESC = 500
LEN_CODE = 50
LEN_COLOR = 20
LEN_DEFAULT = 200
LEN_TEMPLATE_NAME = 100

TRUE_WORDS = {"是", "true", "1", "y", "yes", "required", "√", "✓"}
FALSE_WORDS = {"否", "false", "0", "n", "no", "", "none", "×", "✗"}

HEX_COLOR_RE = re.compile(r"^#?[0-9a-fA-F]{6}$")


class RowError(Exception):
    """单行校验失败，携带可读原因。"""

    def __init__(self, reason: str):
        super().__init__(reason)
        self.reason = reason


# ─────────────────────────── 工具函数 ───────────────────────────

def _now() -> str:
    return datetime.now().isoformat()


def _new_id() -> str:
    return uuid.uuid4().hex[:12]


def _norm(value: Any) -> str | None:
    """单元格值归一：None/空串 → None；其余去首尾空白。"""
    if value is None:
        return None
    if isinstance(value, float) and value.is_integer():
        value = int(value)
    s = str(value).strip()
    return s or None


def _req(value: Any, label: str, max_len: int | None = None) -> str:
    s = _norm(value)
    if not s:
        raise RowError(f"{label}不能为空")
    if max_len and len(s) > max_len:
        raise RowError(f"{label}超长（{len(s)} > {max_len} 字符）：{s[:30]}...")
    return s


def _opt(value: Any, max_len: int | None = None, label: str = "") -> str:
    s = _norm(value)
    if s is None:
        return ""
    if max_len and len(s) > max_len:
        raise RowError(f"{label or '字段'}超长（{len(s)} > {max_len} 字符）")
    return s


def _to_bool(value: Any, label: str) -> int:
    s = (_norm(value) or "").lower()
    if s in FALSE_WORDS:
        return 0
    if s in TRUE_WORDS:
        return 1
    raise RowError(f"{label}取值非法「{value}」，允许：是 / 否 / true / false / 1 / 0")


def _to_int(value: Any, label: str, default: int = 0) -> int:
    s = _norm(value)
    if s is None:
        return default
    try:
        return int(float(s))
    except (TypeError, ValueError):
        raise RowError(f"{label}必须是整数，实际为「{value}」")


def _to_data_type(value: Any) -> str:
    s = (_norm(value) or "").lower()
    if not s:
        raise RowError("数据类型不能为空")
    if s not in DATA_TYPES:
        raise RowError(
            f"数据类型「{value}」不合法，允许值：{' / '.join(sorted(DATA_TYPES))}"
        )
    return s


def _to_color(value: Any, group: str | None, seq: int) -> str | None:
    s = _norm(value)
    if not s:
        if group and group in GROUP_COLORS:
            return GROUP_COLORS[group]
        return PALETTE[seq % len(PALETTE)]
    if not HEX_COLOR_RE.match(s):
        raise RowError(f"颜色「{value}」不是合法的 6 位十六进制色值（如 #5470c6）")
    return s if s.startswith("#") else f"#{s}"


def _header_index(row: tuple) -> dict[str, int]:
    """把表头行解析成 {规范化列名: 列下标}。兼容 '*' 标记与空格差异。"""
    idx: dict[str, int] = {}
    for i, cell in enumerate(row):
        name = _norm(cell)
        if not name:
            continue
        key = name.replace("*", "").replace(" ", "").strip()
        if key and key not in idx:
            idx[key] = i
    return idx


def _cell(row: tuple, idx: dict[str, int], key: str) -> Any:
    i = idx.get(key)
    if i is None or i >= len(row):
        return None
    return row[i]


# ─────────────────────────── 报告结构 ───────────────────────────

class Report:
    """导入结果统计。"""

    def __init__(self) -> None:
        self.sheets: dict[str, dict[str, int]] = {}
        self.errors: list[dict[str, Any]] = []
        self.warnings: list[dict[str, Any]] = []

    def ensure(self, sheet: str) -> dict[str, int]:
        return self.sheets.setdefault(
            sheet, {"total": 0, "success": 0, "failed": 0, "created": 0, "updated": 0, "skipped": 0}
        )

    def ok(self, sheet: str, action: str) -> None:
        st = self.ensure(sheet)
        st["total"] += 1
        st["success"] += 1
        st[action] = st.get(action, 0) + 1

    def fail(self, sheet: str, excel_row: int, reason: str, key: str = "") -> None:
        st = self.ensure(sheet)
        st["total"] += 1
        st["failed"] += 1
        self.errors.append({
            "sheet": sheet,
            "row": excel_row,
            "key": key,
            "reason": reason,
        })

    def warn(self, sheet: str, excel_row: int, message: str) -> None:
        self.warnings.append({"sheet": sheet, "row": excel_row, "message": message})

    def to_dict(self) -> dict[str, Any]:
        total = sum(s["total"] for s in self.sheets.values())
        success = sum(s["success"] for s in self.sheets.values())
        failed = sum(s["failed"] for s in self.sheets.values())
        created = sum(s["created"] for s in self.sheets.values())
        updated = sum(s["updated"] for s in self.sheets.values())
        skipped = sum(s["skipped"] for s in self.sheets.values())
        return {
            "total": total,
            "success": success,
            "failed": failed,
            "created": created,
            "updated": updated,
            "skipped": skipped,
            "sheets": self.sheets,
            "errors": self.errors,
            "warnings": self.warnings,
        }


# ─────────────────────────── 模板 / 导出 ───────────────────────────

HEADER_FILL = PatternFill("solid", fgColor="E8EEF7")
HEADER_FONT = Font(bold=True, size=11)


def _write_sheet(wb: Workbook, sheet_name: str, headers: list[str], rows: list[list[Any]]) -> None:
    ws = wb.create_sheet(sheet_name)
    ws.append(headers)
    for c in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
    for r in rows:
        ws.append(r)
    # 列宽
    for c in range(1, len(headers) + 1):
        width = max(len(str(headers[c - 1])) * 2 + 4, 12)
        for r in rows[:60]:
            v = "" if c > len(r) or r[c - 1] is None else str(r[c - 1])
            width = max(width, min(len(v) * 2 + 4, 52))
        ws.column_dimensions[get_column_letter(c)].width = min(width, 52)
    ws.freeze_panes = "A2"


def build_template_workbook(scope: str = "full", with_example: bool = True) -> io.BytesIO:
    """生成空模板（含可选示例行）。

    Args:
        scope: 模板范围：full / ontologies / relations / constraints / templates
        with_example: 是否附带示例行
    """
    if scope not in VALID_SCOPES:
        raise ValueError(f"scope 必须是 {VALID_SCOPES} 之一")

    wb = Workbook()
    wb.remove(wb.active)

    sheets = SCOPE_SHEETS[scope]

    def _maybe(sheet_name: str):
        return sheet_name if sheet_name in sheets else None

    if with_example:
        cat_rows = [["民航维修领域本体", "面向机务维修与故障根因分析的领域本体（示例，可删除）", "否"]]
        ont_rows = [
            ["民航维修领域本体", "部件", "航空器结构类", "LRU / SRU / 时寿件", "#5470c6", 1],
            ["民航维修领域本体", "故障模式", "故障类", "现象层，可观测的异常表现", "#ee6666", 2],
        ]
        attr_rows = [
            ["民航维修领域本体", "部件", "件号PN", "part_number", "string", "是", "如 65-90751-4", "", 1],
            ["民航维修领域本体", "故障模式", "ATA章节", "ata_chapter", "string", "否", "如 29", "", 2],
        ]
        rel_rows = [["民航维修领域本体", "发生于", "occurs_on", "故障模式/原因 → 部件（双树缝合点）"]]
        con_rows = [["民航维修领域本体", "故障模式", "发生于", "部件", "双树缝合点"]]
        tpl_rows = [["件号基础属性", "件号级实体通用属性"]]
        tpl_attr_rows = [["件号基础属性", "件号PN", "part_number", "string", "是", "", "", 1]]
        bind_rows = [["民航维修领域本体", "部件", "件号基础属性"]]
    else:
        cat_rows, ont_rows, attr_rows, rel_rows = [], [], [], []
        con_rows, tpl_rows, tpl_attr_rows, bind_rows = [], [], [], []

    if _maybe(SHEET_CATEGORY):
        _write_sheet(wb, SHEET_CATEGORY, HEADERS[SHEET_CATEGORY], cat_rows)
    if _maybe(SHEET_ONTOLOGY):
        _write_sheet(wb, SHEET_ONTOLOGY, HEADERS[SHEET_ONTOLOGY], ont_rows)
    if _maybe(SHEET_ATTRIBUTE):
        _write_sheet(wb, SHEET_ATTRIBUTE, HEADERS[SHEET_ATTRIBUTE], attr_rows)
    if _maybe(SHEET_RELATION):
        _write_sheet(wb, SHEET_RELATION, HEADERS[SHEET_RELATION], rel_rows)
    if _maybe(SHEET_CONSTRAINT):
        _write_sheet(wb, SHEET_CONSTRAINT, HEADERS[SHEET_CONSTRAINT], con_rows)
    if _maybe(SHEET_TEMPLATE):
        _write_sheet(wb, SHEET_TEMPLATE, HEADERS[SHEET_TEMPLATE], tpl_rows)
    if _maybe(SHEET_TEMPLATE_ATTR):
        _write_sheet(wb, SHEET_TEMPLATE_ATTR, HEADERS[SHEET_TEMPLATE_ATTR], tpl_attr_rows)
    if _maybe(SHEET_BINDING):
        _write_sheet(wb, SHEET_BINDING, HEADERS[SHEET_BINDING], bind_rows)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


async def export_workbook(db: AsyncSession, category_id: str | None = None, scope: str = "full") -> io.BytesIO:
    """把库中已有本体数据按 scope 导出为 Excel，便于修改后重新导入。"""
    if scope not in VALID_SCOPES:
        raise ValueError(f"scope 必须是 {VALID_SCOPES} 之一")

    sheets = set(SCOPE_SHEETS[scope])

    cat_q = select(OntologyCategory)
    if category_id:
        cat_q = cat_q.where(OntologyCategory.id == category_id)
    categories = (await db.execute(cat_q)).scalars().all()

    cat_rows, ont_rows, attr_rows = [], [], []
    rel_rows, con_rows, bind_rows = [], [], []
    tpl_rows, tpl_attr_rows = [], []

    if scope in ("full", "ontologies"):
        for cat in categories:
            cat_rows.append([cat.name, cat.description or "", "是" if cat.is_system else "否"])

            onts = (await db.execute(
                select(Ontology).where(Ontology.category_id == cat.id).order_by(Ontology.sort_order)
            )).scalars().all()

            for o in onts:
                ont_rows.append([cat.name, o.name, "", o.description or "", o.color or "", o.sort_order])

            if onts:
                ont_ids = [o.id for o in onts]
                ont_name_by_id = {o.id: o.name for o in onts}

                attrs = (await db.execute(
                    select(OntologyAttribute)
                    .where(OntologyAttribute.ontology_id.in_(ont_ids))
                    .order_by(OntologyAttribute.ontology_id, OntologyAttribute.sort_order)
                )).scalars().all()
                for a in attrs:
                    attr_rows.append([
                        cat.name, ont_name_by_id.get(a.ontology_id, ""), a.name, a.code or "",
                        a.data_type, "是" if a.is_required else "否", a.description or "",
                        a.default_value or "", a.sort_order,
                    ])

                binds = (await db.execute(
                    select(OntologyTemplateBinding)
                    .where(OntologyTemplateBinding.ontology_id.in_(ont_ids))
                )).scalars().all()
                if binds:
                    tpl_ids = sorted({b.template_id for b in binds})
                    tpls = (await db.execute(
                        select(OntologyAttributeTemplate)
                        .where(OntologyAttributeTemplate.id.in_(tpl_ids))
                    )).scalars().all()
                    tpl_name = {t.id: t.name for t in tpls}
                    for b in binds:
                        bind_rows.append([
                            cat.name,
                            ont_name_by_id.get(b.ontology_id, ""),
                            tpl_name.get(b.template_id, "")
                        ])

    if scope in ("full", "relations"):
        for cat in categories:
            rels = (await db.execute(
                select(OntologyRelation).where(OntologyRelation.category_id == cat.id)
            )).scalars().all()
            for r in rels:
                rel_rows.append([cat.name, r.name, r.code or "", r.description or ""])

    if scope in ("full", "constraints"):
        for cat in categories:
            rels = (await db.execute(
                select(OntologyRelation).where(OntologyRelation.category_id == cat.id)
            )).scalars().all()
            rel_name = {r.id: r.name for r in rels}

            cons = (await db.execute(
                select(OntologyRelationConstraint)
                .where(OntologyRelationConstraint.category_id == cat.id)
            )).scalars().all()
            if cons:
                ref_ids = set()
                for c in cons:
                    ref_ids.update([c.source_ontology_id, c.target_ontology_id])
                refs = (await db.execute(
                    select(Ontology).where(Ontology.id.in_(list(ref_ids)))
                )).scalars().all()
                ref_name = {o.id: o.name for o in refs}
                for c in cons:
                    con_rows.append([
                        cat.name,
                        ref_name.get(c.source_ontology_id, ""),
                        rel_name.get(c.relation_id, ""),
                        ref_name.get(c.target_ontology_id, ""),
                        c.description or "",
                    ])

    if scope in ("full", "templates"):
        # 模板全局，不受类别过滤
        templates = (await db.execute(
            select(OntologyAttributeTemplate)
        )).scalars().all()
        for t in templates:
            tpl_rows.append([t.name, t.description or ""])
            tas = (await db.execute(
                select(OntologyTemplateAttribute)
                .where(OntologyTemplateAttribute.template_id == t.id)
                .order_by(OntologyTemplateAttribute.sort_order)
            )).scalars().all()
            for ta in tas:
                tpl_attr_rows.append([
                    t.name, ta.name, ta.code or "", ta.data_type,
                    "是" if ta.is_required else "否", ta.description or "",
                    ta.default_value or "", ta.sort_order,
                ])

    wb = Workbook()
    wb.remove(wb.active)
    if SHEET_CATEGORY in sheets:
        _write_sheet(wb, SHEET_CATEGORY, HEADERS[SHEET_CATEGORY], cat_rows)
    if SHEET_ONTOLOGY in sheets:
        _write_sheet(wb, SHEET_ONTOLOGY, HEADERS[SHEET_ONTOLOGY], ont_rows)
    if SHEET_ATTRIBUTE in sheets:
        _write_sheet(wb, SHEET_ATTRIBUTE, HEADERS[SHEET_ATTRIBUTE], attr_rows)
    if SHEET_RELATION in sheets:
        _write_sheet(wb, SHEET_RELATION, HEADERS[SHEET_RELATION], rel_rows)
    if SHEET_CONSTRAINT in sheets:
        _write_sheet(wb, SHEET_CONSTRAINT, HEADERS[SHEET_CONSTRAINT], con_rows)
    if SHEET_TEMPLATE in sheets:
        _write_sheet(wb, SHEET_TEMPLATE, HEADERS[SHEET_TEMPLATE], tpl_rows)
    if SHEET_TEMPLATE_ATTR in sheets:
        _write_sheet(wb, SHEET_TEMPLATE_ATTR, HEADERS[SHEET_TEMPLATE_ATTR], tpl_attr_rows)
    if SHEET_BINDING in sheets:
        _write_sheet(wb, SHEET_BINDING, HEADERS[SHEET_BINDING], bind_rows)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ─────────────────────────── 导入 ───────────────────────────

async def import_workbook(db: AsyncSession, content: bytes, dry_run: bool = False,
                          scope: str = "full") -> dict[str, Any]:
    """解析并执行导入。

    Args:
        db: 数据库会话
        content: xlsx 文件字节
        dry_run: 仅校验不写入（用于「先检查再导入」）
        scope: 导入范围：full / ontologies / relations / constraints / templates

    Returns:
        结构化报告，见 Report.to_dict()
    """
    from openpyxl import load_workbook

    if scope not in VALID_SCOPES:
        return {
            "total": 0, "success": 0, "failed": 0, "created": 0, "updated": 0, "skipped": 0,
            "sheets": {}, "warnings": [],
            "errors": [{"sheet": "-", "row": 0, "key": "", "reason": f"scope 非法：{scope}"}],
            "fatal": f"scope 必须是 {VALID_SCOPES} 之一",
        }

    try:
        wb = load_workbook(io.BytesIO(content), data_only=True, read_only=True)
    except Exception as e:  # noqa: BLE001
        return {
            "total": 0, "success": 0, "failed": 0, "created": 0, "updated": 0, "skipped": 0,
            "sheets": {}, "warnings": [],
            "errors": [{"sheet": "-", "row": 0, "key": "", "reason": f"无法解析 Excel 文件：{e}"}],
            "fatal": f"无法解析 Excel 文件：{e}",
        }

    report = Report()
    sheet_names = set(wb.sheetnames)
    allowed_sheets = set(SCOPE_SHEETS[scope])

    # ── 1. 本体类别 ──
    cat_id_by_name: dict[str, str] = {}
    declared: list[str] = []

    if SHEET_CATEGORY in sheet_names and SHEET_CATEGORY in allowed_sheets:
        ws = wb[SHEET_CATEGORY]
        rows = list(ws.iter_rows(values_only=True))
        if rows:
            idx = _header_index(rows[0])
            for r_no, row in enumerate(rows[1:], start=2):
                if not any(_norm(c) for c in row):
                    continue
                try:
                    name = _req(_cell(row, idx, "类别名称"), "类别名称", LEN_NAME)
                    desc = _opt(_cell(row, idx, "描述"), LEN_DESC, "描述")
                    is_system = _to_bool(_cell(row, idx, "是否系统内置"), "是否系统内置") \
                        if _norm(_cell(row, idx, "是否系统内置")) else 0
                except RowError as e:
                    report.fail(SHEET_CATEGORY, r_no, e.reason, _norm(_cell(row, idx, "类别名称")) or "")
                    continue

                existing = (await db.execute(
                    select(OntologyCategory).where(OntologyCategory.name == name)
                )).scalar_one_or_none()

                if existing:
                    if existing.is_system and not is_system:
                        report.warn(SHEET_CATEGORY, r_no,
                                    f"类别「{name}」为系统内置，已跳过系统标记修改")
                    existing.description = desc
                    existing.updated_at = _now()
                    cat_id_by_name[name] = existing.id
                    report.ok(SHEET_CATEGORY, "updated")
                else:
                    cat = OntologyCategory(
                        id=_new_id(), name=name, description=desc,
                        is_system=int(is_system),
                    )
                    db.add(cat)
                    await db.flush()
                    cat_id_by_name[name] = cat.id
                    report.ok(SHEET_CATEGORY, "created")
                declared.append(name)

    default_cat: str | None = declared[0] if len(declared) == 1 else None

    async def resolve_cat(sheet: str, r_no: int, raw: Any) -> str | None:
        """解析行所属的类别 id；返回 None 表示该行应跳过（已记错误）。"""
        name = _norm(raw)
        if not name:
            if default_cat:
                return cat_id_by_name[default_cat]
            report.fail(sheet, r_no, "未指定类别名称（请在「类别名称」列填写，或在「本体类别」sheet 中定义类别）", "")
            return None
        if name in cat_id_by_name:
            return cat_id_by_name[name]
        # 子表引用了未在「本体类别」sheet 声明的类别 → 自动创建
        existing = (await db.execute(
            select(OntologyCategory).where(OntologyCategory.name == name)
        )).scalar_one_or_none()
        if existing:
            cat_id_by_name[name] = existing.id
        else:
            cat = OntologyCategory(id=_new_id(), name=name, description="")
            db.add(cat)
            await db.flush()
            cat_id_by_name[name] = cat.id
            report.warn(sheet, r_no, f"类别「{name}」未在「{SHEET_CATEGORY}」表中声明，已自动创建")
        return cat_id_by_name[name]

    # ── 2. 本体 ──
    # (cat_id, name) -> id
    ont_id: dict[tuple[str, str], str] = {}
    if SHEET_ONTOLOGY in sheet_names and SHEET_ONTOLOGY in allowed_sheets:
        ws = wb[SHEET_ONTOLOGY]
        rows = list(ws.iter_rows(values_only=True))
        if rows:
            idx = _header_index(rows[0])
            for seq, (r_no, row) in enumerate(
                [(i, r) for i, r in enumerate(rows[1:], start=2) if any(_norm(c) for c in r)]
            ):
                try:
                    cat_id = await resolve_cat(SHEET_ONTOLOGY, r_no, _cell(row, idx, "类别名称"))
                    if not cat_id:
                        continue
                    name = _req(_cell(row, idx, "本体名称"), "本体名称", LEN_NAME)
                    group = _norm(_cell(row, idx, "分组"))
                    desc = _opt(_cell(row, idx, "描述"), LEN_DESC, "描述")
                    color = _to_color(_cell(row, idx, "颜色"), group, seq)
                    sort_order = _to_int(_cell(row, idx, "排序"), "排序", seq)
                    key = (cat_id, name)
                    if key in ont_id:
                        report.warn(SHEET_ONTOLOGY, r_no, f"本体「{name}」在文件中重复定义，以最后一行为准")
                except RowError as e:
                    report.fail(SHEET_ONTOLOGY, r_no, e.reason, _norm(_cell(row, idx, "本体名称")) or "")
                    continue

                existing = (await db.execute(
                    select(Ontology).where(
                        Ontology.category_id == cat_id, Ontology.name == name
                    )
                )).scalar_one_or_none()

                if existing:
                    existing.description = desc
                    existing.color = color
                    existing.sort_order = sort_order
                    existing.updated_at = _now()
                    ont_id[key] = existing.id
                    report.ok(SHEET_ONTOLOGY, "updated")
                else:
                    o = Ontology(
                        id=_new_id(), category_id=cat_id, name=name,
                        description=desc, color=color, sort_order=sort_order,
                    )
                    db.add(o)
                    await db.flush()
                    ont_id[key] = o.id
                    report.ok(SHEET_ONTOLOGY, "created")

    # ── 3. 属性 ──
    if SHEET_ATTRIBUTE in sheet_names and SHEET_ATTRIBUTE in allowed_sheets:
        ws = wb[SHEET_ATTRIBUTE]
        rows = list(ws.iter_rows(values_only=True))
        if rows:
            idx = _header_index(rows[0])
            for r_no, row in enumerate(rows[1:], start=2):
                if not any(_norm(c) for c in row):
                    continue
                ont_name = _norm(_cell(row, idx, "本体名称"))
                attr_name = _norm(_cell(row, idx, "属性名"))
                try:
                    cat_id = await resolve_cat(SHEET_ATTRIBUTE, r_no, _cell(row, idx, "类别名称"))
                    if not cat_id:
                        continue
                    ont_name = _req(_cell(row, idx, "本体名称"), "本体名称", LEN_NAME)
                    attr_name = _req(_cell(row, idx, "属性名"), "属性名", LEN_NAME)
                    code = _opt(_cell(row, idx, "编码"), LEN_CODE, "编码")
                    data_type = _to_data_type(_cell(row, idx, "数据类型"))
                    required = _to_bool(_cell(row, idx, "必填"), "必填")
                    desc = _opt(_cell(row, idx, "描述"), LEN_DESC, "描述")
                    default = _opt(_cell(row, idx, "默认值"), LEN_DEFAULT, "默认值")
                    sort_order = _to_int(_cell(row, idx, "排序"), "排序", 0)

                    key = (cat_id, ont_name)
                    oid = ont_id.get(key)
                    if not oid:
                        exist_o = (await db.execute(
                            select(Ontology).where(
                                Ontology.category_id == cat_id, Ontology.name == ont_name
                            )
                        )).scalar_one_or_none()
                        if not exist_o:
                            raise RowError(f"本体「{ont_name}」不存在（请先在「{SHEET_ONTOLOGY}」表中定义）")
                        oid = exist_o.id
                        ont_id[key] = oid
                except RowError as e:
                    report.fail(SHEET_ATTRIBUTE, r_no, e.reason, f"{ont_name}.{attr_name}")
                    continue

                existing = (await db.execute(
                    select(OntologyAttribute).where(
                        OntologyAttribute.ontology_id == oid,
                        OntologyAttribute.name == attr_name,
                    )
                )).scalar_one_or_none()

                if existing:
                    existing.code = code or None
                    existing.data_type = data_type
                    existing.description = desc
                    existing.is_required = required
                    existing.default_value = default or None
                    existing.sort_order = sort_order
                    existing.updated_at = _now()
                    report.ok(SHEET_ATTRIBUTE, "updated")
                else:
                    db.add(OntologyAttribute(
                        id=_new_id(), ontology_id=oid, name=attr_name,
                        code=code or None, data_type=data_type, description=desc,
                        is_required=required, default_value=default or None,
                        sort_order=sort_order,
                    ))
                    report.ok(SHEET_ATTRIBUTE, "created")

    # ── 4. 关系 ──
    rel_id: dict[tuple[str, str], str] = {}
    if SHEET_RELATION in sheet_names and SHEET_RELATION in allowed_sheets:
        ws = wb[SHEET_RELATION]
        rows = list(ws.iter_rows(values_only=True))
        if rows:
            idx = _header_index(rows[0])
            for r_no, row in enumerate(rows[1:], start=2):
                if not any(_norm(c) for c in row):
                    continue
                rel_name = _norm(_cell(row, idx, "关系名称"))
                try:
                    cat_id = await resolve_cat(SHEET_RELATION, r_no, _cell(row, idx, "类别名称"))
                    if not cat_id:
                        continue
                    rel_name = _req(_cell(row, idx, "关系名称"), "关系名称", LEN_NAME)
                    code = _opt(_cell(row, idx, "编码"), LEN_CODE, "编码")
                    desc = _opt(_cell(row, idx, "描述"), LEN_DESC, "描述")
                except RowError as e:
                    report.fail(SHEET_RELATION, r_no, e.reason, rel_name or "")
                    continue

                existing = (await db.execute(
                    select(OntologyRelation).where(
                        OntologyRelation.category_id == cat_id,
                        OntologyRelation.name == rel_name,
                    )
                )).scalar_one_or_none()
                if existing:
                    existing.code = code or None
                    existing.description = desc
                    existing.updated_at = _now()
                    rel_id[(cat_id, rel_name)] = existing.id
                    report.ok(SHEET_RELATION, "updated")
                else:
                    r = OntologyRelation(
                        id=_new_id(), category_id=cat_id, name=rel_name,
                        code=code or None, description=desc,
                    )
                    db.add(r)
                    await db.flush()
                    rel_id[(cat_id, rel_name)] = r.id
                    report.ok(SHEET_RELATION, "created")

    # ── 5. 约束 ──
    if SHEET_CONSTRAINT in sheet_names and SHEET_CONSTRAINT in allowed_sheets:
        ws = wb[SHEET_CONSTRAINT]
        rows = list(ws.iter_rows(values_only=True))
        if rows:
            idx = _header_index(rows[0])
            for r_no, row in enumerate(rows[1:], start=2):
                if not any(_norm(c) for c in row):
                    continue
                src = _norm(_cell(row, idx, "起点本体"))
                rel = _norm(_cell(row, idx, "关系"))
                tgt = _norm(_cell(row, idx, "终点本体"))
                try:
                    cat_id = await resolve_cat(SHEET_CONSTRAINT, r_no, _cell(row, idx, "类别名称"))
                    if not cat_id:
                        continue
                    src = _req(_cell(row, idx, "起点本体"), "起点本体", LEN_NAME)
                    rel = _req(_cell(row, idx, "关系"), "关系", LEN_NAME)
                    tgt = _req(_cell(row, idx, "终点本体"), "终点本体", LEN_NAME)
                    desc = _opt(_cell(row, idx, "描述"), LEN_DESC, "描述")

                    async def _ont_id(nm: str) -> str:
                        key = (cat_id, nm)
                        if key in ont_id:
                            return ont_id[key]
                        o = (await db.execute(
                            select(Ontology).where(
                                Ontology.category_id == cat_id, Ontology.name == nm
                            )
                        )).scalar_one_or_none()
                        if not o:
                            raise RowError(f"本体「{nm}」不存在（请先在「{SHEET_ONTOLOGY}」表中定义）")
                        ont_id[key] = o.id
                        return o.id

                    async def _rel_id(nm: str) -> str:
                        key = (cat_id, nm)
                        if key in rel_id:
                            return rel_id[key]
                        r = (await db.execute(
                            select(OntologyRelation).where(
                                OntologyRelation.category_id == cat_id,
                                OntologyRelation.name == nm,
                            )
                        )).scalar_one_or_none()
                        if not r:
                            raise RowError(f"关系「{nm}」不存在（请先在「{SHEET_RELATION}」表中定义）")
                        rel_id[key] = r.id
                        return r.id

                    sid = await _ont_id(src)
                    rid = await _rel_id(rel)
                    tid = await _ont_id(tgt)
                except RowError as e:
                    report.fail(SHEET_CONSTRAINT, r_no, e.reason, f"{src}-{rel}->{tgt}")
                    continue

                existing = (await db.execute(
                    select(OntologyRelationConstraint).where(
                        OntologyRelationConstraint.category_id == cat_id,
                        OntologyRelationConstraint.source_ontology_id == sid,
                        OntologyRelationConstraint.relation_id == rid,
                        OntologyRelationConstraint.target_ontology_id == tid,
                    )
                )).scalar_one_or_none()
                if existing:
                    existing.description = desc
                    report.ok(SHEET_CONSTRAINT, "skipped")
                else:
                    db.add(OntologyRelationConstraint(
                        id=_new_id(), category_id=cat_id,
                        source_ontology_id=sid, relation_id=rid,
                        target_ontology_id=tid, description=desc,
                    ))
                    report.ok(SHEET_CONSTRAINT, "created")

    # ── 6. 属性模板 ──
    tpl_id_by_name: dict[str, str] = {}
    if SHEET_TEMPLATE in sheet_names and SHEET_TEMPLATE in allowed_sheets:
        ws = wb[SHEET_TEMPLATE]
        rows = list(ws.iter_rows(values_only=True))
        if rows:
            idx = _header_index(rows[0])
            for r_no, row in enumerate(rows[1:], start=2):
                if not any(_norm(c) for c in row):
                    continue
                tname = _norm(_cell(row, idx, "模板名"))
                try:
                    tname = _req(_cell(row, idx, "模板名"), "模板名", LEN_TEMPLATE_NAME)
                    tdesc = _opt(_cell(row, idx, "模板描述"), LEN_DESC, "模板描述")
                except RowError as e:
                    report.fail(SHEET_TEMPLATE, r_no, e.reason, tname or "")
                    continue

                existing = (await db.execute(
                    select(OntologyAttributeTemplate)
                    .where(OntologyAttributeTemplate.name == tname)
                )).scalar_one_or_none()
                if existing:
                    existing.description = tdesc
                    existing.updated_at = _now()
                    tpl_id_by_name[tname] = existing.id
                    report.ok(SHEET_TEMPLATE, "updated")
                else:
                    t = OntologyAttributeTemplate(
                        id=_new_id(), name=tname, description=tdesc,
                    )
                    db.add(t)
                    await db.flush()
                    tpl_id_by_name[tname] = t.id
                    report.ok(SHEET_TEMPLATE, "created")

    # ── 7. 模板属性 ──
    if SHEET_TEMPLATE_ATTR in sheet_names and SHEET_TEMPLATE_ATTR in allowed_sheets:
        ws = wb[SHEET_TEMPLATE_ATTR]
        rows = list(ws.iter_rows(values_only=True))
        if rows:
            idx = _header_index(rows[0])
            for r_no, row in enumerate(rows[1:], start=2):
                if not any(_norm(c) for c in row):
                    continue
                tname = _norm(_cell(row, idx, "模板名"))
                aname = _norm(_cell(row, idx, "属性名"))
                try:
                    tname = _req(_cell(row, idx, "模板名"), "模板名", LEN_TEMPLATE_NAME)
                    aname = _req(_cell(row, idx, "属性名"), "属性名", LEN_NAME)
                    code = _opt(_cell(row, idx, "编码"), LEN_CODE, "编码")
                    data_type = _to_data_type(_cell(row, idx, "数据类型"))
                    required = _to_bool(_cell(row, idx, "必填"), "必填")
                    desc = _opt(_cell(row, idx, "描述"), LEN_DESC, "描述")
                    default = _opt(_cell(row, idx, "默认值"), LEN_DEFAULT, "默认值")
                    sort_order = _to_int(_cell(row, idx, "排序"), "排序", 0)

                    tid = tpl_id_by_name.get(tname)
                    if not tid:
                        t = (await db.execute(
                            select(OntologyAttributeTemplate)
                            .where(OntologyAttributeTemplate.name == tname)
                        )).scalar_one_or_none()
                        if not t:
                            raise RowError(f"属性模板「{tname}」不存在（请先在「{SHEET_TEMPLATE}」表中定义）")
                        tid = t.id
                        tpl_id_by_name[tname] = tid
                except RowError as e:
                    report.fail(SHEET_TEMPLATE_ATTR, r_no, e.reason, f"{tname}.{aname}")
                    continue

                existing = (await db.execute(
                    select(OntologyTemplateAttribute).where(
                        OntologyTemplateAttribute.template_id == tid,
                        OntologyTemplateAttribute.name == aname,
                    )
                )).scalar_one_or_none()
                if existing:
                    existing.code = code or None
                    existing.data_type = data_type
                    existing.description = desc
                    existing.is_required = required
                    existing.default_value = default or None
                    existing.sort_order = sort_order
                    existing.updated_at = _now()
                    report.ok(SHEET_TEMPLATE_ATTR, "updated")
                else:
                    db.add(OntologyTemplateAttribute(
                        id=_new_id(), template_id=tid, name=aname,
                        code=code or None, data_type=data_type, description=desc,
                        is_required=required, default_value=default or None,
                        sort_order=sort_order,
                    ))
                    report.ok(SHEET_TEMPLATE_ATTR, "created")

    # ── 8. 模板绑定 ──
    if SHEET_BINDING in sheet_names and SHEET_BINDING in allowed_sheets:
        ws = wb[SHEET_BINDING]
        rows = list(ws.iter_rows(values_only=True))
        if rows:
            idx = _header_index(rows[0])
            for r_no, row in enumerate(rows[1:], start=2):
                if not any(_norm(c) for c in row):
                    continue
                oname = _norm(_cell(row, idx, "本体名称"))
                tname = _norm(_cell(row, idx, "模板名"))
                try:
                    cat_id = await resolve_cat(SHEET_BINDING, r_no, _cell(row, idx, "类别名称"))
                    if not cat_id:
                        continue
                    oname = _req(_cell(row, idx, "本体名称"), "本体名称", LEN_NAME)
                    tname = _req(_cell(row, idx, "模板名"), "模板名", LEN_TEMPLATE_NAME)

                    key = (cat_id, oname)
                    oid = ont_id.get(key)
                    if not oid:
                        o = (await db.execute(
                            select(Ontology).where(
                                Ontology.category_id == cat_id, Ontology.name == oname
                            )
                        )).scalar_one_or_none()
                        if not o:
                            raise RowError(f"本体「{oname}」不存在")
                        oid = o.id
                        ont_id[key] = oid

                    tid = tpl_id_by_name.get(tname)
                    if not tid:
                        t = (await db.execute(
                            select(OntologyAttributeTemplate)
                            .where(OntologyAttributeTemplate.name == tname)
                        )).scalar_one_or_none()
                        if not t:
                            raise RowError(f"属性模板「{tname}」不存在")
                        tid = t.id
                        tpl_id_by_name[tname] = tid
                except RowError as e:
                    report.fail(SHEET_BINDING, r_no, e.reason, f"{oname}<-{tname}")
                    continue

                existing = (await db.execute(
                    select(OntologyTemplateBinding).where(
                        OntologyTemplateBinding.ontology_id == oid,
                        OntologyTemplateBinding.template_id == tid,
                    )
                )).scalar_one_or_none()
                if existing:
                    report.ok(SHEET_BINDING, "skipped")
                else:
                    db.add(OntologyTemplateBinding(
                        id=_new_id(), ontology_id=oid, template_id=tid, sort_order=0,
                    ))
                    report.ok(SHEET_BINDING, "created")

    wb.close()

    if dry_run:
        await db.rollback()
        result = report.to_dict()
        result["dry_run"] = True
        return result

    await db.commit()
    result = report.to_dict()
    result["dry_run"] = False
    return result
